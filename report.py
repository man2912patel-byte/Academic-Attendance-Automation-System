import os
import logging
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

logger = logging.getLogger("report")

class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically calculate the total page count
    and render 'Page X of Y' in the footer.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#555555"))
        
        # Header (Top of each page except the first maybe, but we can do all)
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.5)
        self.line(36, 756, 576, 756)
        self.drawString(36, 762, "MFT Academic Attendance Automation System - Summary Report")
        
        # Footer
        self.line(36, 54, 576, 54)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(576, 40, page_text)
        self.drawString(36, 40, f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}")
        self.restoreState()


class ReportGenerator:
    def __init__(self, config, history_mgr):
        self.config = config
        self.history_mgr = history_mgr

    def generate_daily_pdf(self, date_str, output_path=None):
        """
        Generates a styled PDF report for a single date.
        """
        run = self.history_mgr.get_run_by_date(date_str)
        if not run:
            raise ValueError(f"No attendance data found in database for date: {date_str}")
            
        details = self.history_mgr.get_details_by_date(date_str)
        if not details:
            raise ValueError(f"No student details found for date: {date_str}")

        if output_path is None:
            output_path = Path(self.config.reports_dir) / f"Daily_Report_{date_str}.pdf"
        else:
            output_path = Path(output_path)

        # Initialize PDF Document
        # Letter is 612 x 792 points. Margins: 36 points (0.5 inch)
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=54,
            bottomMargin=54
        )

        styles = getSampleStyleSheet()
        
        # Custom Paragraph styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1E3D59"),
            alignment=1, # Center
            spaceAfter=15
        )
        
        normal_style = ParagraphStyle(
            'ReportBody',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#333333")
        )

        bold_style = ParagraphStyle(
            'ReportBold',
            parent=normal_style,
            fontName='Helvetica-Bold'
        )

        story = []
        
        # Title
        story.append(Paragraph(f"DAILY ACADEMIC ATTENDANCE REPORT", title_style))
        story.append(Spacer(1, 10))
        
        # Stats summary block
        pct = (run["present_count"] / run["total_students"] * 100) if run["total_students"] > 0 else 0
        stats_data = [
            [Paragraph("Attendance Date:", bold_style), Paragraph(date_str, normal_style), 
             Paragraph("Sync/Generation Time:", bold_style), Paragraph(run["sync_time"], normal_style)],
            [Paragraph("Total Students:", bold_style), Paragraph(str(run["total_students"]), normal_style),
             Paragraph("Present Count:", bold_style), Paragraph(str(run["present_count"]), normal_style)],
            [Paragraph("Absent Count:", bold_style), Paragraph(str(run["absent_count"]), normal_style),
             Paragraph("Attendance %:", bold_style), Paragraph(f"{pct:.2f}%", bold_style)]
        ]
        
        stats_table = Table(stats_data, colWidths=[110, 140, 130, 160])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F5F7FA")),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#D9D9D9")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Students Table
        table_headers = [
            Paragraph("<b>Roll No</b>", bold_style),
            Paragraph("<b>Enrollment No</b>", bold_style),
            Paragraph("<b>Student Name</b>", bold_style),
            Paragraph("<b>Status</b>", bold_style),
            Paragraph("<b>Time Generated</b>", bold_style)
        ]
        
        table_rows = [table_headers]
        
        # Colors for status
        present_style = ParagraphStyle('PresentStyle', parent=bold_style, textColor=colors.HexColor("#385723"))
        absent_style = ParagraphStyle('AbsentStyle', parent=bold_style, textColor=colors.HexColor("#C00000"))
        
        for idx, rec in enumerate(details):
            status_text = rec["attendance"]
            status_para = Paragraph(status_text, present_style if status_text == "Present" else absent_style)
            
            table_rows.append([
                Paragraph(rec["roll_number"], normal_style),
                Paragraph(rec["enrollment_number"], normal_style),
                Paragraph(rec["student_name"], normal_style),
                status_para,
                Paragraph(datetime.datetime.now().strftime("%H:%M:%S"), normal_style)
            ])

        # Widths of columns: Total width 540 points (612 - 72 margins)
        student_table = Table(table_rows, colWidths=[85, 115, 200, 70, 70])
        
        # Styling Student Table
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3D59")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#1E3D59")),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ]
        
        # Let's change the text color of the headers to white in TableStyle explicitly
        for col in range(len(table_headers)):
            t_style.append(('TEXTCOLOR', (col, 0), (col, 0), colors.white))
            
        # Add alternating row background
        for i in range(1, len(table_rows)):
            bg = colors.HexColor("#F9FAFB") if i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
            # Present/Absent background highlight
            status_cell = details[i-1]["attendance"]
            if status_cell == "Present":
                t_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor("#E2F0D9")))
            else:
                t_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor("#FCE4D6")))

        student_table.setStyle(TableStyle(t_style))
        story.append(student_table)
        
        # Build Document with NumberedCanvas for header/footer
        doc.build(story, canvasmaker=NumberedCanvas)
        logger.info(f"Daily PDF report generated at {output_path}")
        return output_path

    def generate_monthly_report(self, year, month, format_type="xlsx", output_dir=None):
        """
        Generates a summary report of attendance for all dates in a specific month.
        Displays each student's attendance on every class date and their monthly present %.
        Formats: 'xlsx' or 'pdf'.
        """
        # Fetch all runs in that month
        all_runs = self.history_mgr.get_all_runs()
        month_str = f"{year:04d}-{month:02d}"
        
        month_dates = []
        for run in all_runs:
            if run["attendance_date"].startswith(month_str):
                month_dates.append(run["attendance_date"])
                
        # Sort dates chronologically
        month_dates = sorted(month_dates)
        
        if not month_dates:
            raise ValueError(f"No attendance data found in history for {month_str}.")

        # Get unique students in the history
        students = self.history_mgr.get_student_list()
        
        # Pull attendance matrix
        # student_key -> date -> attendance
        matrix = {}
        for student in students:
            matrix[student["roll_number"]] = {d: "N/A" for d in month_dates}

        for date_str in month_dates:
            details = self.history_mgr.get_details_by_date(date_str)
            for d in details:
                roll = d["roll_number"]
                if roll in matrix:
                    matrix[roll][date_str] = d["attendance"]

        if output_dir is None:
            if format_type.lower() == "pdf":
                output_dir = Path(self.config.reports_dir)
            else:
                output_dir = Path(self.config.output_dir)
        else:
            output_dir = Path(output_dir)
            
        output_dir.mkdir(parents=True, exist_ok=True)
        month_name = datetime.date(year, month, 1).strftime("%B_%Y")
        
        if format_type.lower() == "xlsx":
            return self._generate_monthly_excel(students, month_dates, matrix, month_name, month_str, output_dir)
        elif format_type.lower() == "pdf":
            return self._generate_monthly_pdf(students, month_dates, matrix, month_name, month_str, output_dir)
        else:
            raise ValueError(f"Unsupported report format: {format_type}")

    def _generate_monthly_excel(self, students, dates, matrix, month_name, month_str, output_dir):
        filename = output_dir / f"Monthly_Attendance_Report_{month_name}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        ws.views.sheetView[0].showGridLines = True
        
        font_name = "Segoe UI"
        title_font = Font(name=font_name, size=15, bold=True, color="1E3D59")
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=11)
        meta_font = Font(name=font_name, size=10, italic=True)
        
        header_fill = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
        zebra_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        present_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        present_font = Font(name=font_name, size=10, bold=True, color="385723")
        
        absent_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        absent_font = Font(name=font_name, size=10, bold=True, color="C00000")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Merge Title
        total_cols = 4 + len(dates)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(row=1, column=1, value=f"MONTHLY ACADEMIC ATTENDANCE SHEET - {month_name.replace('_', ' ')}").font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Headers Row
        headers = ["Roll No", "Enrollment No", "Student Name"]
        # Format date headers to display DD/MM
        date_headers = []
        for d in dates:
            dt = datetime.datetime.strptime(d, "%Y-%m-%d")
            date_headers.append(dt.strftime("%d-%b"))
            
        headers.extend(date_headers)
        headers.append("Present %")
        
        for col_num, val in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 30
        
        # Write Data
        for r_idx, s in enumerate(students, 4):
            roll = s["roll_number"]
            ws.cell(row=r_idx, column=1, value=roll)
            ws.cell(row=r_idx, column=2, value=s["enrollment_number"])
            ws.cell(row=r_idx, column=3, value=s["student_name"])
            
            p_count = 0
            valid_days = 0
            
            # Write attendance for each date
            for c_idx, date_str in enumerate(dates, 4):
                status = matrix[roll].get(date_str, "N/A")
                cell = ws.cell(row=r_idx, column=c_idx)
                
                if status == "Present":
                    cell.value = "P"
                    cell.fill = present_fill
                    cell.font = present_font
                    p_count += 1
                    valid_days += 1
                elif status == "Absent":
                    cell.value = "A"
                    cell.fill = absent_fill
                    cell.font = absent_font
                    valid_days += 1
                else:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal="center")
                    
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Present %
            pct = (p_count / valid_days * 100) if valid_days > 0 else 0.0
            pct_cell = ws.cell(row=r_idx, column=4 + len(dates), value=f"{pct:.1f}%")
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.border = thin_border
            
            # Formatting % color
            if pct >= 75.0:
                pct_cell.font = Font(name=font_name, size=11, bold=True, color="385723")
            else:
                pct_cell.font = Font(name=font_name, size=11, bold=True, color="C00000")
                
            # Zebra row striping for first 3 static columns
            is_zebra = (r_idx % 2 == 0)
            for c in range(1, 4):
                cell = ws.cell(row=r_idx, column=c)
                cell.font = data_font
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill
                    
            ws.row_dimensions[r_idx].height = 22
            
        # Adjust Column Widths
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column in [1, 2]:
                ws.column_dimensions[col_letter].width = 15
            elif col[0].column == 3:
                ws.column_dimensions[col_letter].width = 25
            elif col[0].column <= 3 + len(dates):
                ws.column_dimensions[col_letter].width = 8 # date headers (P/A cells)
            else:
                ws.column_dimensions[col_letter].width = 12 # % column
                
        wb.save(filename)
        logger.info(f"Monthly Excel report saved to {filename}")
        return filename

    def _generate_monthly_pdf(self, students, dates, matrix, month_name, month_str, output_dir):
        """
        Generates a landscape summary report of monthly attendance in PDF.
        """
        filename = output_dir / f"Monthly_Attendance_Report_{month_name}.pdf"
        
        # Use landscape orientation since monthly sheets are wide
        doc = SimpleDocTemplate(
            str(filename),
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=40,
            bottomMargin=40
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'MonthlyTitle',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor("#1E3D59"),
            alignment=1, # Center
            spaceAfter=15
        )
        
        normal_style = ParagraphStyle('MonthlyNormal', fontName='Helvetica', fontSize=7, leading=9)
        bold_style = ParagraphStyle('MonthlyBold', fontName='Helvetica-Bold', fontSize=7, leading=9)
        
        story = []
        story.append(Paragraph(f"MONTHLY ACADEMIC ATTENDANCE SUMMARY - {month_name.replace('_', ' ')}", title_style))
        
        # We need to construct a table
        headers = [Paragraph("<b>Roll No</b>", bold_style), Paragraph("<b>Student Name</b>", bold_style)]
        for d in dates:
            dt = datetime.datetime.strptime(d, "%Y-%m-%d")
            headers.append(Paragraph(f"<b>{dt.strftime('%d/%m')}</b>", bold_style))
        headers.append(Paragraph("<b>Pres %</b>", bold_style))
        
        table_rows = [headers]
        
        for student in students:
            roll = student["roll_number"]
            row = [
                Paragraph(roll, normal_style),
                Paragraph(student["student_name"][:18], normal_style) # Truncate long names to fit
            ]
            
            p_count = 0
            valid_days = 0
            for d in dates:
                status = matrix[roll].get(d, "N/A")
                if status == "Present":
                    row.append(Paragraph("P", normal_style))
                    p_count += 1
                    valid_days += 1
                elif status == "Absent":
                    row.append(Paragraph("A", normal_style))
                    valid_days += 1
                else:
                    row.append(Paragraph("-", normal_style))
                    
            pct = (p_count / valid_days * 100) if valid_days > 0 else 0.0
            row.append(Paragraph(f"<b>{pct:.1f}%</b>", bold_style))
            table_rows.append(row)

        # Width calculation:
        # Total landscape letter width is 792 points. Margins: 36 left/right = 720 points printable.
        # Let's allocate: Roll no = 60, Student name = 110, % = 45.
        # Remaining width is 720 - 215 = 505 points. Divide this by number of dates.
        col_count = len(dates)
        date_col_width = min(max(505 / col_count, 15), 45)
        
        col_widths = [60, 110] + [date_col_width] * col_count + [45]
        
        t = Table(table_rows, colWidths=col_widths)
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3D59")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (1,0), (1,-1), 'LEFT'), # Left align name
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]
        
        # Color codes in cells
        for r_idx in range(1, len(table_rows)):
            bg = colors.HexColor("#F9FAFB") if r_idx % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), bg))
            
            # Col 2 to col_count + 2
            roll = students[r_idx - 1]["roll_number"]
            for c_idx in range(2, len(dates) + 2):
                d = dates[c_idx - 2]
                status = matrix[roll].get(d, "N/A")
                if status == "Present":
                    t_style.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#E2F0D9")))
                    t_style.append(('TEXTCOLOR', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#385723")))
                elif status == "Absent":
                    t_style.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#FCE4D6")))
                    t_style.append(('TEXTCOLOR', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#C00000")))

        t.setStyle(TableStyle(t_style))
        story.append(t)
        
        doc.build(story)
        logger.info(f"Monthly PDF report generated at {filename}")
        return filename
