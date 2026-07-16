import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from models import db, AttendanceRun, StudentAttendance

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#555555"))
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.5)
        
        self.line(36, 576, 756, 576)
        self.drawString(36, 582, "Academic Attendance Automation System - Summary Report")
        
        self.line(36, 45, 756, 45)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(756, 32, page_text)
        self.drawString(36, 32, f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}")
        self.restoreState()


def compile_runs_list(user_id, start_date=None, end_date=None, year=None, month=None):
    """Fetches all runs based on filters using SQLAlchemy ORM."""
    query = AttendanceRun.query.filter_by(user_id=user_id)
    
    if start_date and end_date:
        query = query.filter(AttendanceRun.attendance_date.between(start_date, end_date))
    elif year and month:
        start = datetime.date(year, month, 1)
        if month == 12:
            end = datetime.date(year + 1, 1, 1)
        else:
            end = datetime.date(year, month + 1, 1)
        query = query.filter(AttendanceRun.attendance_date >= start, AttendanceRun.attendance_date < end)
        
    return query.order_by(AttendanceRun.attendance_date.asc()).all()


def get_month_students(user_id, run_ids):
    """Fetches list of unique students under given run ids using SQLAlchemy."""
    if not run_ids:
        return []
        
    records = StudentAttendance.query.filter(StudentAttendance.run_id.in_(run_ids)).order_by(StudentAttendance.roll_number.asc()).all()
    
    seen = set()
    unique_students = []
    for r in records:
        if r.roll_number not in seen:
            seen.add(r.roll_number)
            unique_students.append({
                "roll_number": r.roll_number,
                "enrollment_number": r.enrollment_number or "N/A",
                "student_name": r.student_name or "Unknown Student"
            })
    return unique_students


def build_matrix_report_excel(user_id, title_text, runs):
    """Generates styled Excel binary stream for monthly/custom matrix reports using SQLAlchemy."""
    run_ids = [r.id for r in runs]
    dates = [r.attendance_date.strftime("%Y-%m-%d") for r in runs]
    students = get_month_students(user_id, run_ids)
    
    matrix = {}
    for s in students:
        matrix[s["roll_number"]] = {d: "N/A" for d in dates}
        
    # Populate matrix
    for r in runs:
        date_str = r.attendance_date.strftime("%Y-%m-%d")
        for record in r.records:
            if record.roll_number in matrix:
                matrix[record.roll_number][date_str] = record.attendance

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary Report"
    ws.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Arial", size=14, bold=True, color="1E3D59")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    bold_data_font = Font(name="Arial", size=9, bold=True)
    
    header_fill = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
    present_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    absent_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Title Row
    ws.merge_cells("A1:E1")
    ws["A1"] = title_text
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Headers
    headers = ["Roll No", "Enrollment No", "Student Name"] + dates + ["Present Days", "Attendance %"]
    ws.append([]) # spacer
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border
        
    # Data rows
    current_row = 4
    for idx, s in enumerate(students):
        roll = s["roll_number"]
        row_data = [roll, s["enrollment_number"], s["student_name"]]
        
        present_count = 0
        total_days = len(dates)
        
        for d in dates:
            status = matrix[roll][d]
            row_data.append(status)
            if status == "Present":
                present_count += 1
                
        rate = (present_count / total_days * 100) if total_days > 0 else 0
        row_data.append(present_count)
        row_data.append(f"{round(rate, 1)}%")
        
        ws.append(row_data)
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = border
            
            if col_idx <= 3:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
            val = str(cell.value)
            if val == "Present":
                cell.fill = present_fill
            elif val == "Absent":
                cell.fill = absent_fill
                
        ws.cell(row=current_row, column=len(row_data)).font = bold_data_font
        current_row += 1
        
    # Auto-fit widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col[0].column <= 3:
            for cell in col[2:]:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        else:
            ws.column_dimensions[col_letter].width = 11

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_matrix_report_pdf(user_id, title_text, runs):
    """Generates styled PDF binary stream for monthly/custom matrix reports using SQLAlchemy."""
    run_ids = [r.id for r in runs]
    dates = [r.attendance_date.strftime("%Y-%m-%d") for r in runs]
    students = get_month_students(user_id, run_ids)
    
    matrix = {}
    for s in students:
        matrix[s["roll_number"]] = {d: "N/A" for d in dates}
        
    for r in runs:
        date_str = r.attendance_date.strftime("%Y-%m-%d")
        for record in r.records:
            if record.roll_number in matrix:
                matrix[record.roll_number][date_str] = record.attendance

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=36, leftMargin=36, topMargin=54, bottomMargin=54)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=colors.HexColor('#1E3D59'),
        spaceAfter=15
    )
    
    th_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=7,
        textColor=colors.white,
        alignment=1
    )
    
    td_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=7,
        alignment=1
    )
    
    td_left_style = ParagraphStyle(
        'TableCellLeft',
        fontName='Helvetica',
        fontSize=7,
        alignment=0
    )

    story = []
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 10))
    
    headers = ["Roll No", "Student Name"] + [d[5:] for d in dates] + ["Pres", "%"]
    
    col_widths = [55, 110]
    date_col_width = max(18, min(40, 420 / max(1, len(dates))))
    for _ in dates:
        col_widths.append(date_col_width)
    col_widths.append(25)
    col_widths.append(30)
    
    table_data = [[Paragraph(h, th_style) for h in headers]]
    
    for s in students:
        roll = s["roll_number"]
        name = s["student_name"]
        
        row = [
            Paragraph(roll, td_left_style),
            Paragraph(name[:20], td_left_style)
        ]
        
        present_count = 0
        total_days = len(dates)
        
        for d in dates:
            status = matrix[roll][d]
            if status == "Present":
                present_count += 1
                status_char = "P"
            elif status == "Absent":
                status_char = "A"
            else:
                status_char = "-"
            row.append(Paragraph(status_char, td_style))
            
        rate = (present_count / total_days * 100) if total_days > 0 else 0
        row.append(Paragraph(str(present_count), td_style))
        row.append(Paragraph(f"{int(rate)}%", td_style))
        
        table_data.append(row)
        
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3D59')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]
    
    for row_idx in range(1, len(table_data)):
        for col_idx in range(2, len(col_widths) - 2):
            cell_para = table_data[row_idx][col_idx]
            if cell_para.text == "P":
                t_style.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#D1E7DD')))
                t_style.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#0F5132')))
            elif cell_para.text == "A":
                t_style.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#F8D7DA')))
                t_style.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#842029')))
                
    table.setStyle(TableStyle(t_style))
    story.append(table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer
