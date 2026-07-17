import urllib.request
import csv
import sys
import io

# Set CSV field size limit globally to prevent OverflowError or Field Limit Errors
max_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(max_limit)
        break
    except OverflowError:
        max_limit = int(max_limit / 10)

import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from pathlib import Path
from flask import Blueprint, request, jsonify, send_file, current_app
from utils.auth import token_required
from models import db, AttendanceRun, StudentAttendance
from utils.attendance_processor import (
    parse_sheets_data,
    match_students,
    compute_attendance
)

attendance_bp = Blueprint('attendance', __name__)

import os
import tempfile

import time
import traceback

import re

def convert_gsheet_url(url):
    if not url or not isinstance(url, str):
        return url
    if "docs.google.com/spreadsheets" in url and "/edit" in url:
        match_id = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
        if match_id:
            sheet_id = match_id.group(1)
            match_gid = re.search(r'gid=([0-9]+)', url)
            gid = match_gid.group(1) if match_gid else "0"
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    return url

def mask_sheet_url(url):
    if not url:
        return ""
    if "/d/" in url:
        try:
            parts = url.split("/d/")
            subparts = parts[1].split("/")
            sheet_id = subparts[0]
            masked_id = sheet_id[:8] + "..." + sheet_id[-4:] if len(sheet_id) > 12 else "..."
            subparts[0] = masked_id
            parts[1] = "/".join(subparts)
            return "/d/".join(parts)
        except Exception:
            return "Masked URL (Format unresolvable)"
    return url

def fetch_and_parse_sheets(current_user):
    start_time = time.time()
    user_id = current_user.id
    student_src = current_user.student_source_type or 'google_sheet'
    attendance_src = current_user.attendance_source_type or 'google_sheet'
    
    # Initialize log variables
    student_url_masked = mask_sheet_url(convert_gsheet_url(current_user.student_excel_path)) if student_src == 'google_sheet' else "N/A"
    attendance_url_masked = mask_sheet_url(convert_gsheet_url(current_user.attendance_excel_path)) if attendance_src == 'google_sheet' else "N/A"
    student_filename = current_user.student_uploaded_file.split('/')[-1] if (student_src == 'upload' and current_user.student_uploaded_file) else "N/A"
    attendance_filename = current_user.attendance_uploaded_file.split('/')[-1] if (attendance_src == 'upload' and current_user.attendance_uploaded_file) else "N/A"
    
    student_http_status = None
    attendance_http_status = None
    student_bytes_len = 0
    attendance_bytes_len = 0
    student_encoding = "UTF-8"
    attendance_encoding = "UTF-8"
    student_rows = 0
    attendance_rows = 0
    student_cols = 0
    attendance_cols = 0
    student_headers = []
    attendance_headers = []
    error_trace = None

    try:
        # 1. Validate Student configuration
        if student_src == 'google_sheet':
            if not current_user.student_excel_path:
                raise ValueError("No Student List source configured.")
        elif student_src == 'upload':
            if not current_user.student_uploaded_file:
                raise ValueError("No Student List source configured.")
        else:
            raise ValueError("No Student List source configured.")

        # 2. Validate Attendance configuration
        if attendance_src == 'google_sheet':
            if not current_user.attendance_excel_path:
                raise ValueError("No Attendance source configured.")
        elif attendance_src == 'upload':
            if not current_user.attendance_uploaded_file:
                raise ValueError("No Attendance source configured.")
        else:
            raise ValueError("No Attendance source configured.")

        mft_raw = []
        marquee_raw = []

        # 3. Load Student Data
        if student_src == 'upload':
            from services.storage_service import StorageService
            storage = StorageService.get_provider()
            try:
                student_bytes = storage.read_file(current_user.student_uploaded_file)
                student_bytes_len = len(student_bytes)
                ext = os.path.splitext(current_user.student_uploaded_file)[1].lower()
                if ext == '.csv':
                    student_csv = student_bytes.decode('utf-8', errors='ignore')
                    mft_reader = csv.DictReader(student_csv.splitlines())
                    mft_raw = [dict(row) for row in mft_reader]
                else:
                    student_encoding = "Excel Default"
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        tmp.write(student_bytes)
                        tmp_path = tmp.name
                        tmp.close() # Explicitly close the file handle on Windows before reopening it
                    try:
                        from utils.attendance_processor import load_student_list_excel
                        mft_raw = load_student_list_excel(tmp_path)
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                            print(f"[LOG] File path: {tmp_path} - File deleted")
            except Exception as e:
                raise ValueError(f"Uploaded CSV is corrupted. Detail: {str(e)}")
        else:
            student_url = convert_gsheet_url(current_user.student_excel_path)
            print(f"[LOG] Final Student CSV URL: {student_url_masked}")
            student_req = urllib.request.Request(student_url, headers={'User-Agent': 'Mozilla/5.0'})
            try:
                with urllib.request.urlopen(student_req, timeout=15) as r:
                    student_http_status = getattr(r, 'status', 200)
                    student_bytes = r.read()
                student_bytes_len = len(student_bytes)
                student_csv = student_bytes.decode('utf-8', errors='ignore')
                print(f"[LOG] Raw CSV Content from Google Sheets (first 2000 chars):\n{student_csv[:2000]}")
                mft_reader = csv.DictReader(student_csv.splitlines())
                mft_raw = [dict(row) for row in mft_reader]
            except Exception as e:
                raise ValueError(f"Google Sheet cannot be accessed. Detail: {str(e)}")

        student_rows = len(mft_raw)
        student_cols = len(mft_raw[0].keys()) if mft_raw else 0
        student_headers = list(mft_raw[0].keys())[:5] if mft_raw else []

        # 4. Load Attendance Data
        if attendance_src == 'upload':
            from services.storage_service import StorageService
            storage = StorageService.get_provider()
            try:
                attendance_bytes = storage.read_file(current_user.attendance_uploaded_file)
                attendance_bytes_len = len(attendance_bytes)
                ext = os.path.splitext(current_user.attendance_uploaded_file)[1].lower()
                if ext == '.csv':
                    attendance_csv = attendance_bytes.decode('utf-8', errors='ignore')
                    marquee_reader = csv.reader(attendance_csv.splitlines())
                    marquee_raw = [row for row in marquee_reader]
                else:
                    attendance_encoding = "Excel Default"
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        tmp.write(attendance_bytes)
                        tmp_path = tmp.name
                        tmp.close() # Explicitly close the file handle on Windows before reopening it
                    try:
                        from utils.attendance_processor import load_attendance_logs_excel
                        marquee_raw = load_attendance_logs_excel(tmp_path)
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                            print(f"[LOG] File path: {tmp_path} - File deleted")
            except Exception as e:
                raise ValueError(f"Uploaded CSV is corrupted. Detail: {str(e)}")
        else:
            attendance_url = convert_gsheet_url(current_user.attendance_excel_path)
            print(f"[LOG] Final Attendance CSV URL: {attendance_url_masked}")
            attendance_req = urllib.request.Request(attendance_url, headers={'User-Agent': 'Mozilla/5.0'})
            try:
                with urllib.request.urlopen(attendance_req, timeout=15) as r:
                    attendance_http_status = getattr(r, 'status', 200)
                    attendance_bytes = r.read()
                attendance_bytes_len = len(attendance_bytes)
                attendance_csv = attendance_bytes.decode('utf-8', errors='ignore')
                print(f"[LOG] Raw CSV Content from Google Sheets (first 2000 chars):\n{attendance_csv[:2000]}")
                marquee_reader = csv.reader(attendance_csv.splitlines())
                marquee_raw = [row for row in marquee_reader]
            except Exception as e:
                raise ValueError(f"Google Sheet cannot be accessed. Detail: {str(e)}")

        attendance_rows = len(marquee_raw)
        attendance_cols = len(marquee_raw[0]) if marquee_raw else 0
        attendance_headers = marquee_raw[0][:5] if marquee_raw else []

        res = parse_sheets_data(mft_raw, marquee_raw, student_url=student_url_masked, attendance_url=attendance_url_masked)
        
        # Log successful load parameters
        processing_time = time.time() - start_time
        print(f"""
=== PROCESS LOG [SUCCESS] ===
User ID: {user_id}
Student Source Type: {student_src}
Attendance Source Type: {attendance_src}
Student URL: {student_url_masked}
Attendance URL: {attendance_url_masked}
Student Filename: {student_filename}
Attendance Filename: {attendance_filename}
Student HTTP Status: {student_http_status}
Attendance HTTP Status: {attendance_http_status}
Student Downloaded Bytes: {student_bytes_len}
Attendance Downloaded Bytes: {attendance_bytes_len}
Student Encoding: {student_encoding}
Attendance Encoding: {attendance_encoding}
Student Rows: {student_rows}
Attendance Rows: {attendance_rows}
Student Columns: {student_cols}
Attendance Columns: {attendance_cols}
Student Headers: {student_headers}
Attendance Headers: {attendance_headers}
Processing Time: {processing_time:.4f}s
Error Traceback: None
=============================
""")
        return res

    except Exception as e:
        error_trace = traceback.format_exc()
        processing_time = time.time() - start_time
        print(f"""
=== PROCESS LOG [FAILED] ===
User ID: {user_id}
Student Source Type: {student_src}
Attendance Source Type: {attendance_src}
Student URL: {student_url_masked}
Attendance URL: {attendance_url_masked}
Student Filename: {student_filename}
Attendance Filename: {attendance_filename}
Student HTTP Status: {student_http_status}
Attendance HTTP Status: {attendance_http_status}
Student Downloaded Bytes: {student_bytes_len}
Attendance Downloaded Bytes: {attendance_bytes_len}
Student Encoding: {student_encoding}
Attendance Encoding: {attendance_encoding}
Student Rows: {student_rows}
Attendance Rows: {attendance_rows}
Student Columns: {student_cols}
Attendance Columns: {attendance_cols}
Student Headers: {student_headers}
Attendance Headers: {attendance_headers}
Processing Time: {processing_time:.4f}s
Error Traceback:
{error_trace}
============================
""")
        raise ValueError(f"Spreadsheet parsing failed: {str(e)}")

def handle_parse_exception(e):
    err_msg = str(e)
    err_low = err_msg.lower()
    if "no student" in err_low or "no attendance" in err_low:
        return jsonify({'message': err_msg}), 400
    if "invalid student list structure" in err_low or "invalid attendance logs structure" in err_low or "invalid sheet structure" in err_low:
        return jsonify({'message': err_msg}), 400
    if "fetch google sheet" in err_low or "download" in err_low or "not found" in err_low:
        return jsonify({'message': err_msg}), 404
    if "parsing failed" in err_low or "failed to load" in err_low or "read csv" in err_low or "read excel" in err_low or "format" in err_low or "field larger than field limit" in err_low:
        return jsonify({'message': err_msg}), 422
    return jsonify({'message': f'Unexpected backend error: {err_msg}'}), 500

@attendance_bp.route('/attendance/dates', methods=['GET'])
@token_required
def get_attendance_dates(current_user):
    """Fetches unique attendance dates sorted newest first (descending)."""
    try:
        _, _, date_map = fetch_and_parse_sheets(current_user)
        dates_list = sorted([d.strftime("%Y-%m-%d") for d in date_map.keys()], reverse=True)
        return jsonify({'dates': dates_list}), 200
    except Exception as e:
        return handle_parse_exception(e)

@attendance_bp.route('/attendance/session', methods=['GET'])
@token_required
def get_attendance_sessions(current_user):
    """Fetches available session columns for a given date."""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'message': 'Date parameter is required.'}), 400
    try:
        _, _, date_map = fetch_and_parse_sheets(current_user)
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        sessions = date_map.get(date_obj, [])
        session_names = ["Combined (Any)", "Combined (All)"]
        for _, name in sessions:
            if name not in session_names:
                session_names.append(name)
        return jsonify({'sessions': session_names}), 200
    except Exception as e:
        return handle_parse_exception(e)

@attendance_bp.route('/attendance/upload-data', methods=['GET'])
@token_required
def get_upload_data(current_user):
    """Parses configured sources and returns raw JSON metadata for preview/generation in the frontend."""
    try:
        mft_students, marquee_students, date_map = fetch_and_parse_sheets(current_user)
        
        serialized_date_map = {}
        for d_obj, sessions in date_map.items():
            date_str = d_obj.strftime("%Y-%m-%d")
            serialized_date_map[date_str] = [
                {"colIdx": item[0], "sessionName": item[1]} for item in sessions
            ]
            
        dates_list = sorted([d.strftime("%Y-%m-%d") for d in date_map.keys()], reverse=True)
        
        date_sessions = {}
        for d_str, sessions in serialized_date_map.items():
            date_sessions[d_str] = ["Combined (Any)", "Combined (All)"] + [
                s["sessionName"] for s in sessions
            ]
            
        return jsonify({
            'dates': dates_list,
            'dateSessions': date_sessions,
            'mftStudents': mft_students,
            'marqueeStudents': marquee_students,
            'dateMap': serialized_date_map
        }), 200
    except Exception as e:
        return handle_parse_exception(e)

@attendance_bp.route('/attendance/generate', methods=['POST'])
@token_required
def generate_attendance(current_user):
    """Generates attendance matching, statistics, and stores the run in SQLite."""
    data = request.get_json() or {}
    date_str = data.get('date')
    session_mode = data.get('session_mode') or data.get('session', 'Combined (Any)')
    
    if not date_str:
        return jsonify({'message': 'Date parameter is required.'}), 400
        
    try:
        mft_students, marquee_students, date_map = fetch_and_parse_sheets(current_user)
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        session_cols = date_map.get(date_obj)
        
        if not session_cols:
            return jsonify({'message': f'No session records found for selected date: {date_str}'}), 404
            
        matched_records, mismatched_mft = match_students(mft_students, marquee_students)
        records = compute_attendance(matched_records, mismatched_mft, date_obj, session_cols, session_mode)
        
        total = len(records)
        present = sum(1 for r in records if r["attendance"] == "Present")
        absent = total - present
        rate = round((present / total * 100) if total > 0 else 0, 1)
        
        # Prevent unique constraints violation by deleting existing run for same date
        existing = AttendanceRun.query.filter_by(user_id=current_user.id, attendance_date=date_obj).first()
        if existing:
            db.session.delete(existing)
            db.session.commit()
            
        new_run = AttendanceRun(
            user_id=current_user.id,
            attendance_date=date_obj,
            session_name=session_mode,
            total_students=total,
            present_count=present,
            absent_count=absent,
            attendance_rate=rate,
            excel_file_path=str(Path(current_app.config['EXPORTS_DIR']) / current_user.username / f"Academic_Attendance_{date_str}.xlsx"),
            pdf_file_path=str(Path(current_app.config['EXPORTS_DIR']) / current_user.username / f"Academic_Attendance_{date_str}.pdf")
        )
        db.session.add(new_run)
        db.session.commit()
        
        for r in records:
            new_record = StudentAttendance(
                run_id=new_run.id,
                roll_number=r["roll_number"],
                enrollment_number=r["enrollment_number"],
                student_name=r["student_name"],
                attendance=r["attendance"]
            )
            db.session.add(new_record)
        db.session.commit()
        
        return jsonify({
            'records': records,
            'summary': {
                'total': total,
                'present': present,
                'absent': absent,
                'rate': rate
            }
        }), 200
    except Exception as e:
        return handle_parse_exception(e)

@attendance_bp.route('/attendance/save-run', methods=['POST'])
@token_required
def save_run_compat(current_user):
    """Compat route to save run."""
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    summary = data.get('summary', {})
    if not date_str or not records:
        return jsonify({'message': 'Date and student records are required.'}), 400
    try:
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        existing = AttendanceRun.query.filter_by(user_id=current_user.id, attendance_date=date_obj).first()
        if existing:
            db.session.delete(existing)
            db.session.commit()
            
        tot = int(summary.get('total', len(records)))
        pres = int(summary.get('present', 0))
        pct_rate = round((pres / tot * 100) if tot > 0 else 0.0, 1)
        new_run = AttendanceRun(
            user_id=current_user.id,
            attendance_date=date_obj,
            session_name=summary.get('session_name', 'Combined (Any)'),
            total_students=tot,
            present_count=pres,
            absent_count=int(summary.get('absent', 0)),
            attendance_rate=pct_rate,
            excel_file_path=str(Path(current_app.config['EXPORTS_DIR']) / current_user.username / f"Academic_Attendance_{date_str}.xlsx"),
            pdf_file_path=str(Path(current_app.config['EXPORTS_DIR']) / current_user.username / f"Academic_Attendance_{date_str}.pdf")
        )
        db.session.add(new_run)
        db.session.commit()
        
        for r in records:
            new_record = StudentAttendance(
                run_id=new_run.id,
                roll_number=r["roll_number"],
                enrollment_number=r["enrollment_number"],
                student_name=r["student_name"],
                attendance=r["attendance"]
            )
            db.session.add(new_record)
        db.session.commit()
        return jsonify({'success': True}), 201
    except Exception as e:
        return jsonify({'message': str(e)}), 500

@attendance_bp.route('/history', methods=['GET'])
@attendance_bp.route('/attendance/history', methods=['GET'])
@token_required
def get_history(current_user):
    """Retrieves paginated and searchable history logs."""
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    search = request.args.get('search', '').strip()
    
    try:
        query = AttendanceRun.query.filter_by(user_id=current_user.id)
        if search:
            query = query.filter(AttendanceRun.attendance_date.like(f"%{search}%"))
            
        total_matching = query.count()
        runs_query = query.order_by(AttendanceRun.attendance_date.desc()).paginate(page=page, per_page=limit, error_out=False)
        
        runs = []
        for r in runs_query.items:
            pct = (r.present_count / r.total_students * 100) if r.total_students > 0 else 0
            runs.append({
                "id": r.id,
                "attendance_date": r.attendance_date.strftime("%Y-%m-%d"),
                "sync_time": r.sync_time.strftime("%Y-%m-%d %H:%M:%S"),
                "total_students": r.total_students,
                "present_count": r.present_count,
                "absent_count": r.absent_count,
                "rate": round(pct, 1)
            })
            
        total_pages = (total_matching + limit - 1) // limit if total_matching > 0 else 1
        return jsonify({
            "runs": runs,
            "total_pages": total_pages,
            "total_runs": total_matching,
            "current_page": page
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve history: {str(e)}'}), 500

@attendance_bp.route('/history/<int:run_id>', methods=['GET'])
@attendance_bp.route('/attendance/history/<int:run_id>', methods=['GET'])
@token_required
def get_run_details(current_user, run_id):
    """Retrieves details of an attendance run."""
    try:
        run = AttendanceRun.query.filter_by(user_id=current_user.id, id=run_id).first()
        if not run:
            return jsonify({'message': 'Run not found.'}), 404
            
        search_val = request.args.get('search', '').strip().lower()
        status_val = request.args.get('status', '').strip().lower()
        
        query = StudentAttendance.query.filter_by(run_id=run_id)
        if status_val:
            query = query.filter(db.func.lower(StudentAttendance.attendance) == status_val)
            
        details = []
        for d in query.all():
            roll = d.roll_number or ""
            enroll = d.enrollment_number or ""
            name = d.student_name or ""
            status = d.attendance or ""
            
            if not search_val or (search_val in name.lower() or search_val in roll.lower() or search_val in enroll.lower()):
                details.append({
                    "id": d.id,
                    "roll_number": roll,
                    "enrollment_number": enroll,
                    "student_name": name,
                    "attendance": status
                })
                
        pct = (run.present_count / run.total_students * 100) if run.total_students > 0 else 0
        return jsonify({
            "run": {
                "id": run.id,
                "attendance_date": run.attendance_date.strftime("%Y-%m-%d"),
                "total_students": run.total_students,
                "present_count": run.present_count,
                "absent_count": run.absent_count,
                "rate": round(pct, 1)
            },
            "details": details
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve details: {str(e)}'}), 500

@attendance_bp.route('/history/<int:run_id>', methods=['DELETE'])
@attendance_bp.route('/attendance/history/<int:run_id>', methods=['DELETE'])
@token_required
def delete_run(current_user, run_id):
    """Deletes an attendance run and cascades to all detailed student marks."""
    try:
        run = AttendanceRun.query.filter_by(user_id=current_user.id, id=run_id).first()
        if not run:
            return jsonify({'message': 'Run not found.'}), 404
        db.session.delete(run)
        db.session.commit()
        return jsonify({'message': 'Attendance run deleted successfully.'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to delete run: {str(e)}'}), 500

@attendance_bp.route('/reports/stats', methods=['GET'])
@attendance_bp.route('/attendance/reports/stats', methods=['GET'])
@token_required
def get_reports_stats(current_user):
    """Retrieves options to configure reports generation."""
    try:
        runs = AttendanceRun.query.filter_by(user_id=current_user.id).order_by(AttendanceRun.attendance_date.desc()).all()
        dates = [r.attendance_date.strftime("%Y-%m-%d") for r in runs]
        
        months_seen = set()
        months_options = []
        
        total_runs = len(runs)
        total_present = sum(r.present_count for r in runs)
        total_students = sum(r.total_students for r in runs)
        
        for r in runs:
            dt = r.attendance_date
            year = dt.year
            month = dt.month
            month_label = dt.strftime("%B %Y")
            month_key = f"{year}-{month}"
            
            if month_key not in months_seen:
                months_seen.add(month_key)
                months_options.append({
                    "year": year,
                    "month": month,
                    "label": month_label
                })
                
        avg_rate = round((total_present / total_students * 100) if total_students > 0 else 0, 1)
        
        return jsonify({
            "dates": dates,
            "months": months_options,
            "summary": {
                "total_runs": total_runs,
                "average_rate": avg_rate
            }
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve report metadata: {str(e)}'}), 500

@attendance_bp.route('/reports', methods=['POST'])
@attendance_bp.route('/reports/generate', methods=['POST'])
@attendance_bp.route('/attendance/reports/generate', methods=['POST'])
@token_required
def generate_report_file(current_user):
    """Compiles daily, monthly, or custom matrix reports as Excel or PDF downloads."""
    from utils.report_generator import compile_runs_list, build_matrix_report_excel, build_matrix_report_pdf

    data = request.get_json() or {}
    report_type = data.get('report_type', 'daily')
    fmt = data.get('format', 'excel')
    
    try:
        if report_type == 'daily':
            date_str = data.get('date')
            if not date_str:
                return jsonify({'message': 'Date parameter is required for Daily reports.'}), 400
                
            run = AttendanceRun.query.filter_by(user_id=current_user.id).filter(AttendanceRun.attendance_date == datetime.datetime.strptime(date_str, "%Y-%m-%d").date()).first()
            if not run:
                return jsonify({'message': f'No attendance run found for date {date_str}.'}), 404
                
            records = [{"roll_number": d.roll_number, "enrollment_number": d.enrollment_number, "student_name": d.student_name, "attendance": d.attendance} for d in run.records]
            
            if fmt == 'pdf':
                return export_pdf_file(date_str, records)
            else:
                return export_excel_file(date_str, records)
                
        elif report_type == 'monthly':
            year = data.get('year')
            month = data.get('month')
            if not year or not month:
                return jsonify({'message': 'Year and Month parameters are required.'}), 400
                
            runs = compile_runs_list(current_user.id, year=int(year), month=int(month))
            if not runs:
                return jsonify({'message': 'No attendance history records found for the selected month.'}), 404
                
            month_name = datetime.date(int(year), int(month), 1).strftime("%B %Y")
            title_text = f"MONTHLY ATTENDANCE REPORT - {month_name.upper()}"
            
            if fmt == 'pdf':
                file_stream = build_matrix_report_pdf(current_user.id, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Monthly_Attendance_Report_{year}_{month}.pdf",
                    mimetype="application/pdf"
                )
            else:
                file_stream = build_matrix_report_excel(current_user.id, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Monthly_Attendance_Report_{year}_{month}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        elif report_type == 'custom':
            start_date_str = data.get('start_date')
            end_date_str = data.get('end_date')
            if not start_date_str or not end_date_str:
                return jsonify({'message': 'Start and End dates are required.'}), 400
                
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            
            runs = compile_runs_list(current_user.id, start_date=start_date, end_date=end_date)
            if not runs:
                return jsonify({'message': 'No attendance history records found for the specified range.'}), 404
                
            title_text = f"ATTENDANCE SUMMARY REPORT ({start_date_str} TO {end_date_str})"
            
            if fmt == 'pdf':
                file_stream = build_matrix_report_pdf(current_user.id, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Attendance_Report_{start_date_str}_to_{end_date_str}.pdf",
                    mimetype="application/pdf"
                )
            else:
                file_stream = build_matrix_report_excel(current_user.id, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Attendance_Report_{start_date_str}_to_{end_date_str}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            return jsonify({'message': 'Unsupported report type.'}), 400
    except Exception as e:
        return jsonify({'message': f'Report generation failed: {str(e)}'}), 500

@attendance_bp.route('/attendance/export-excel', methods=['POST'])
@token_required
def export_excel_route(current_user):
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    return export_excel_file(date_str, records)

@attendance_bp.route('/attendance/export-pdf', methods=['POST'])
@token_required
def export_pdf_route(current_user):
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    return export_pdf_file(date_str, records)

def export_excel_file(date_str, records):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Academic Attendance"
        ws.views.sheetView[0].showGridLines = True
        
        font_name = "Segoe UI"
        title_font = Font(name=font_name, size=16, bold=True, color="1E3D59")
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=11)
        
        header_fill = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
        present_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        absent_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        ws.merge_cells("A1:D1")
        ws["A1"] = f"ACADEMIC ATTENDANCE AUTOMATION - {date_str}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ["Roll No", "Enrollment No", "Student Name", "Attendance"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        for r_idx, r in enumerate(records, 4):
            ws.cell(row=r_idx, column=1, value=r["roll_number"]).font = data_font
            ws.cell(row=r_idx, column=2, value=r["enrollment_number"]).font = data_font
            ws.cell(row=r_idx, column=3, value=r["student_name"]).font = data_font
            
            status_cell = ws.cell(row=r_idx, column=4, value=r["attendance"])
            status_cell.font = data_font
            if r["attendance"] == "Present":
                status_cell.fill = present_fill
            else:
                status_cell.fill = absent_fill
                
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return send_file(
            file_stream,
            as_attachment=True,
            download_name=f"Academic_Attendance_{date_str}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({'message': f'Failed to generate Excel: {str(e)}'}), 500

def export_pdf_file(date_str, records):
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='TitleStyle',
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#1E3D59'),
            alignment=1,
            spaceAfter=15
        )
        
        story = []
        story.append(Paragraph(f"Academic Attendance Report - {date_str}", title_style))
        story.append(Spacer(1, 10))
        
        table_data = [["Roll No", "Enrollment No", "Student Name", "Attendance"]]
        for r in records:
            table_data.append([
                r["roll_number"],
                r["enrollment_number"],
                r["student_name"],
                r["attendance"]
            ])
            
        t = Table(table_data, colWidths=[80, 100, 240, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3D59')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ]))
        
        for i, row in enumerate(records, 1):
            color = colors.HexColor('#E2F0D9') if row["attendance"] == "Present" else colors.HexColor('#FCE4D6')
            t.setStyle(TableStyle([
                ('BACKGROUND', (3, i), (3, i), color),
                ('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#385723') if row["attendance"] == "Present" else colors.HexColor('#C00000')),
                ('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'),
            ]))
            
        story.append(t)
        doc.build(story)
        
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Academic_Attendance_{date_str}.pdf",
            mimetype="application/pdf"
        )
    except Exception as e:
        return jsonify({'message': f'Failed to generate PDF: {str(e)}'}), 500
