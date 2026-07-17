import urllib.request
import csv
import sys
import os
import uuid
from datetime import datetime

# Set CSV field size limit globally to prevent OverflowError or Field Limit Errors
max_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(max_limit)
        break
    except OverflowError:
        max_limit = int(max_limit / 10)

from flask import Blueprint, request, jsonify, current_app
from utils.auth import token_required
from models import db, User
from utils.attendance_processor import parse_sheets_data
import openpyxl

settings_bp = Blueprint('settings', __name__)

def validate_student_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    headers = []
    if ext == '.csv':
        try:
            print(f"[LOG] File path: {file_path} - Open started")
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            print(f"[LOG] File path: {file_path} - Read completed")
            reader = csv.reader(content.splitlines())
            headers = next(reader)
        except Exception as e:
            raise ValueError(f"Failed to read CSV: {str(e)}")
    else:
        try:
            print(f"[LOG] File path: {file_path} - Open started")
            with open(file_path, "rb") as f:
                wb = openpyxl.load_workbook(f, read_only=True)
                sheet = wb.active
                rows = sheet.iter_rows(max_row=1, values_only=True)
                headers = next(rows, [])
                wb.close()
            print(f"[LOG] File path: {file_path} - Read completed")
            print(f"[LOG] File path: {file_path} - Workbook closed")
        except Exception as e:
            raise ValueError(f"Failed to read Excel workbook: {str(e)}")
            
    email_ok = False
    roll_ok = False
    for h in headers:
        if h is None:
            continue
        h_low = str(h).strip().lower()
        if "email" in h_low or "mail" in h_low:
            email_ok = True
        elif "rhn" in h_low or "roll" in h_low or "id" in h_low:
            roll_ok = True
            
    if not email_ok and not roll_ok:
        raise ValueError("Invalid Student List structure. File must contain a header for 'Email' or 'Roll Number/ID'.")
        
    if ext == '.csv':
        rows_count = len(content.splitlines()) - 1
        cols_count = len(headers)
    else:
        try:
            print(f"[LOG] File path: {file_path} - Open started")
            with open(file_path, "rb") as f:
                wb = openpyxl.load_workbook(f, data_only=True)
                sheet = wb.active
                rows_count = sheet.max_row - 1 if sheet.max_row > 1 else 0
                cols_count = sheet.max_column
                wb.close()
            print(f"[LOG] File path: {file_path} - Read completed")
            print(f"[LOG] File path: {file_path} - Workbook closed")
        except Exception as e:
            raise ValueError(f"Failed to read Excel workbook metadata: {str(e)}")
        
    return rows_count, cols_count, "CSV" if ext == '.csv' else "Excel"

def validate_attendance_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    rows = []
    if ext == '.csv':
        try:
            print(f"[LOG] File path: {file_path} - Open started")
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            print(f"[LOG] File path: {file_path} - Read completed")
            reader = csv.reader(content.splitlines())
            rows = list(reader)
        except Exception as e:
            raise ValueError(f"Failed to read CSV: {str(e)}")
    else:
        try:
            print(f"[LOG] File path: {file_path} - Open started")
            with open(file_path, "rb") as f:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                sheet = wb.active
                for r in sheet.rows:
                    rows.append([cell.value for cell in r])
                wb.close()
            print(f"[LOG] File path: {file_path} - Read completed")
            print(f"[LOG] File path: {file_path} - Workbook closed")
        except Exception as e:
            raise ValueError(f"Failed to read Excel workbook: {str(e)}")
            
    if not rows or len(rows) < 2:
        raise ValueError("Invalid Attendance Logs structure. File must have at least 2 header rows.")
        
    row0 = [str(cell or "").strip().lower() for cell in rows[0]]
    
    enroll_ok = any("enrollment" in col or "enroll" in col for col in row0)
    name_ok = any("name" in col or "student" in col for col in row0)
    email_ok = any("mail" in col or "email" in col for col in row0)
    
    if not enroll_ok and not name_ok and not email_ok:
        raise ValueError("Invalid Attendance Logs structure. File must contain an 'Enrollment', 'Name', or 'Email' column.")
        
    rows_count = len(rows) - 2 if len(rows) > 2 else 0
    cols_count = len(row0)
    
    return rows_count, cols_count, "CSV" if ext == '.csv' else "Excel"

@settings_bp.route('/settings', methods=['GET'])
@token_required
def get_settings(current_user):
    """Retrieves settings preferences for the logged-in user."""
    try:
        return jsonify(current_user.to_dict()), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve settings: {str(e)}'}), 500

@settings_bp.route('/settings', methods=['PUT'])
@token_required
def update_settings(current_user):
    """Updates settings preferences for the logged-in user."""
    data = request.get_json() or {}
    from services.storage_service import StorageService
    storage = StorageService.get_provider()
    
    if 'student_excel_path' in data:
        new_url = data['student_excel_path'].strip()
        current_user.student_excel_path = new_url
        if new_url:
            current_user.student_source_type = 'google_sheet'
            if current_user.student_uploaded_file:
                storage.delete_file(current_user.student_uploaded_file)
                current_user.student_uploaded_file = None
                current_user.student_file_size = None
                current_user.student_upload_time = None
                current_user.student_rows_detected = None
                current_user.student_columns_detected = None
                current_user.student_file_type = None

    if 'attendance_excel_path' in data:
        new_url = data['attendance_excel_path'].strip()
        current_user.attendance_excel_path = new_url
        if new_url:
            current_user.attendance_source_type = 'google_sheet'
            if current_user.attendance_uploaded_file:
                storage.delete_file(current_user.attendance_uploaded_file)
                current_user.attendance_uploaded_file = None
                current_user.attendance_file_size = None
                current_user.attendance_upload_time = None
                current_user.attendance_rows_detected = None
                current_user.attendance_columns_detected = None
                current_user.attendance_file_type = None

    if 'student_uploaded_file' in data and not data['student_uploaded_file']:
        if current_user.student_uploaded_file:
            storage.delete_file(current_user.student_uploaded_file)
        current_user.student_uploaded_file = None
        current_user.student_source_type = 'google_sheet'
        current_user.student_file_size = None
        current_user.student_upload_time = None
        current_user.student_rows_detected = None
        current_user.student_columns_detected = None
        current_user.student_file_type = None

    if 'attendance_uploaded_file' in data and not data['attendance_uploaded_file']:
        if current_user.attendance_uploaded_file:
            storage.delete_file(current_user.attendance_uploaded_file)
        current_user.attendance_uploaded_file = None
        current_user.attendance_source_type = 'google_sheet'
        current_user.attendance_file_size = None
        current_user.attendance_upload_time = None
        current_user.attendance_rows_detected = None
        current_user.attendance_columns_detected = None
        current_user.attendance_file_type = None

    if 'theme' in data:
        current_user.theme = data['theme'].strip()
    if 'dark_mode' in data:
        current_user.dark_mode = bool(data['dark_mode'])
    if 'output_folder' in data:
        current_user.output_folder = data['output_folder'].strip()
    if 'backup_folder' in data:
        current_user.backup_folder = data['backup_folder'].strip()
    if 'export_format' in data:
        current_user.export_format = data['export_format'].strip()
    if 'auto_backup' in data:
        current_user.auto_backup = bool(data['auto_backup'])
    
    try:
        db.session.commit()
        return jsonify({'message': 'Settings saved successfully.'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to update settings: {str(e)}'}), 500

@settings_bp.route('/settings/upload/student', methods=['POST'])
@token_required
def upload_student(current_user):
    if 'file' not in request.files:
        return jsonify({'message': 'No file part in the request.'}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'message': 'No selected file.'}), 400
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv']:
        return jsonify({'message': 'Unsupported file format. Please upload .xlsx, .xls, or .csv.'}), 400
        
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    
    if size > 20 * 1024 * 1024:
        return jsonify({'message': 'File size exceeds the 20MB limit.'}), 413
        
    unique_filename = f"student_{uuid.uuid4().hex}{ext}"
    from services.storage_service import StorageService
    storage = StorageService.get_provider()
    
    path = storage.save_file(file, current_user.id, unique_filename)
    abs_path = os.path.join(current_app.config['UPLOADS_DIR'], str(current_user.id), unique_filename)
    
    try:
        rows, cols, f_type = validate_student_file(abs_path)
    except Exception as e:
        storage.delete_file(path)
        err_msg = str(e)
        err_low = err_msg.lower()
        if "invalid" in err_low or "structure" in err_low or "must contain" in err_low:
            return jsonify({'message': f'Validation failed: {err_msg}'}), 400
        else:
            return jsonify({'message': f'Parsing failed: {err_msg}'}), 422
        
    if current_user.student_uploaded_file:
        storage.delete_file(current_user.student_uploaded_file)
        
    current_user.student_uploaded_file = path
    current_user.student_source_type = 'upload'
    current_user.student_file_size = size
    current_user.student_upload_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    current_user.student_rows_detected = rows
    current_user.student_columns_detected = cols
    current_user.student_file_type = f_type
    current_user.student_excel_path = None # Clear URL on upload
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'path': path,
        'filename': file.filename,
        'metadata': {
            'file_size': size,
            'upload_time': current_user.student_upload_time,
            'rows_detected': rows,
            'columns_detected': cols,
            'file_type': f_type
        }
    }), 200

@settings_bp.route('/settings/upload/attendance', methods=['POST'])
@token_required
def upload_attendance(current_user):
    if 'file' not in request.files:
        return jsonify({'message': 'No file part in the request.'}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'message': 'No selected file.'}), 400
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv']:
        return jsonify({'message': 'Unsupported file format. Please upload .xlsx, .xls, or .csv.'}), 400
        
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    
    if size > 20 * 1024 * 1024:
        return jsonify({'message': 'File size exceeds the 20MB limit.'}), 413
        
    unique_filename = f"attendance_{uuid.uuid4().hex}{ext}"
    from services.storage_service import StorageService
    storage = StorageService.get_provider()
    
    path = storage.save_file(file, current_user.id, unique_filename)
    abs_path = os.path.join(current_app.config['UPLOADS_DIR'], str(current_user.id), unique_filename)
    
    try:
        rows, cols, f_type = validate_attendance_file(abs_path)
    except Exception as e:
        storage.delete_file(path)
        err_msg = str(e)
        err_low = err_msg.lower()
        if "invalid" in err_low or "structure" in err_low or "must contain" in err_low:
            return jsonify({'message': f'Validation failed: {err_msg}'}), 400
        else:
            return jsonify({'message': f'Parsing failed: {err_msg}'}), 422
        
    if current_user.attendance_uploaded_file:
        storage.delete_file(current_user.attendance_uploaded_file)
        
    current_user.attendance_uploaded_file = path
    current_user.attendance_source_type = 'upload'
    current_user.attendance_file_size = size
    current_user.attendance_upload_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    current_user.attendance_rows_detected = rows
    current_user.attendance_columns_detected = cols
    current_user.attendance_file_type = f_type
    current_user.attendance_excel_path = None # Clear URL on upload
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'path': path,
        'filename': file.filename,
        'metadata': {
            'file_size': size,
            'upload_time': current_user.attendance_upload_time,
            'rows_detected': rows,
            'columns_detected': cols,
            'file_type': f_type
        }
    }), 200

@settings_bp.route('/settings/verify-files', methods=['POST'])
@token_required
def verify_sources(current_user):
    """Verifies that the configured sources exist and parse successfully."""
    student_src = current_user.student_source_type or 'google_sheet'
    attendance_src = current_user.attendance_source_type or 'google_sheet'
    
    if student_src == 'google_sheet' and not current_user.student_excel_path:
        return jsonify({'success': False, 'message': 'Please configure your Student List and Attendance source first.'}), 400
    if student_src == 'upload' and not current_user.student_uploaded_file:
        return jsonify({'success': False, 'message': 'Please configure your Student List and Attendance source first.'}), 400
        
    if attendance_src == 'google_sheet' and not current_user.attendance_excel_path:
        return jsonify({'success': False, 'message': 'Please configure your Student List and Attendance source first.'}), 400
    if attendance_src == 'upload' and not current_user.attendance_uploaded_file:
        return jsonify({'success': False, 'message': 'Please configure your Student List and Attendance source first.'}), 400
        
    try:
        from routes.attendance_routes import fetch_and_parse_sheets
        fetch_and_parse_sheets(current_user)
        return jsonify({
            'success': True,
            'message': 'Both configured sources verified and parsed successfully. Connection status: Active.'
        }), 200
    except Exception as e:
        err_msg = str(e)
        err_low = err_msg.lower()
        if "no student" in err_low or "no attendance" in err_low or "missing" in err_low:
            status_code = 400
        elif "invalid student list structure" in err_low or "invalid attendance logs structure" in err_low or "invalid sheet structure" in err_low:
            status_code = 400
        elif "fetch google sheet" in err_low or "download" in err_low or "not found" in err_low:
            status_code = 404
        elif "parsing failed" in err_low or "failed to load" in err_low or "read csv" in err_low or "read excel" in err_low or "format" in err_low or "field larger than field limit" in err_low:
            status_code = 422
        else:
            status_code = 400
        return jsonify({
            'success': False,
            'message': f'Verification failed: {err_msg}'
        }), status_code

@settings_bp.route('/settings', methods=['DELETE'])
@token_required
def delete_user_data(current_user):
    """Resets user data by deleting associated attendance runs."""
    try:
        from models import AttendanceRun
        AttendanceRun.query.filter_by(user_id=current_user.id).delete()
        db.session.commit()
        return jsonify({'message': 'All user data history deleted successfully.'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to reset user history: {str(e)}'}), 500
