from flask import Blueprint, request, jsonify, send_file, current_app
from app.auth import token_required
from app.utils.db_helpers import get_db_connection
from app.utils.attendance_processor import (
    load_student_list_excel,
    load_attendance_logs_excel,
    parse_sheets_data,
    match_students,
    compute_attendance
)
import datetime
import json
import csv
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from pathlib import Path

attendance_bp = Blueprint('attendance', __name__)

# Cache folder helper
def get_user_history_dir(username):
    path = Path(current_app.config['UPLOADS_DIR']) / username / "history"
    path.mkdir(parents=True, exist_ok=True)
    return path

def get_user_output_dir(username):
    path = Path(current_app.config['EXPORTS_DIR']) / username
    path.mkdir(parents=True, exist_ok=True)
    return path

def get_user_reports_dir(username):
    path = Path(current_app.config['REPORTS_DIR']) / username
    path.mkdir(parents=True, exist_ok=True)
    return path


@attendance_bp.route('/sync', methods=['POST'])
@token_required
def sync_sheets(current_user):
    """Syncs configured local Excel file data and caches locally, returning available dates and sessions."""
    if not current_user.student_excel_path or not current_user.attendance_excel_path:
        return jsonify({'message': 'Please configure local Excel file paths in Settings first.'}), 400
        
    try:
        # Load from configured local Excel files
        mft_raw = load_student_list_excel(current_user.student_excel_path)
        marquee_raw = load_attendance_logs_excel(current_user.attendance_excel_path)
        
        if not mft_raw or not marquee_raw:
            return jsonify({'message': 'Failed to retrieve Excel data. Check that the files are valid and contain sheets.'}), 400
            
        history_dir = get_user_history_dir(current_user.username)
        
        # Save MFT student list cache as CSV
        mft_cache_path = history_dir / "mft_student_cache.csv"
        if mft_raw:
            with open(mft_cache_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=mft_raw[0].keys())
                writer.writeheader()
                writer.writerows(mft_raw)
                
        # Save Marquee raw attendance cache as CSV
        marquee_cache_path = history_dir / "marquee_attendance_cache.csv"
        with open(marquee_cache_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(marquee_raw)
            
        # Parse dates and sessions candidate list
        _, _, date_map = parse_sheets_data(mft_raw, marquee_raw)
        
        # Format response dates and sessions list
        dates_list = []
        date_sessions = {}
        for date_obj, sessions in sorted(date_map.items()):
            date_str = date_obj.strftime("%Y-%m-%d")
            dates_list.append(date_str)
            
            # Prepend Combined modes
            sess_names = ["Combined (Any)", "Combined (All)"]
            seen_sessions = set()
            for _, name in sessions:
                if name not in seen_sessions:
                    sess_names.append(name)
                    seen_sessions.add(name)
            date_sessions[date_str] = sess_names
            
        return jsonify({
            'success': True,
            'dates': dates_list,
            'date_sessions': date_sessions
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Sync failed to read Excel data sources: {str(e)}'}), 500


@attendance_bp.route('/preview', methods=['POST'])
@token_required
def preview_attendance(current_user):
    """Generates preview and returns matches and stats for a selected date/session."""
    data = request.get_json() or {}
    date_str = data.get('date')
    session_mode = data.get('session_mode', 'Combined (Any)')
    
    if not date_str:
        return jsonify({'message': 'Date parameter is required.'}), 400
        
    history_dir = get_user_history_dir(current_user.username)
    mft_cache_path = history_dir / "mft_student_cache.csv"
    marquee_cache_path = history_dir / "marquee_attendance_cache.csv"
    
    if not mft_cache_path.exists() or not marquee_cache_path.exists():
        return jsonify({'message': 'No cached roster files found. Please Load Excel Data Source in Generate page first.'}), 400
        
    try:
        # Load from caches
        mft_raw = []
        with open(mft_cache_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            mft_raw = [dict(row) for row in reader]
            
        marquee_raw = []
        with open(marquee_cache_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            marquee_raw = [row for row in reader]
            
        # Parse, match and compute
        mft_students, marquee_students, date_map = parse_sheets_data(mft_raw, marquee_raw)
        
        # Find session columns
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        session_cols = date_map.get(date_obj)
        
        if not session_cols:
            return jsonify({'message': f'No session records found for selected date: {date_str}'}), 404
            
        matched_records, mismatched_mft = match_students(mft_students, marquee_students)
        records = compute_attendance(matched_records, mismatched_mft, date_obj, session_cols, session_mode)
        
        # Calculate rates
        total = len(records)
        present = sum(1 for r in records if r["attendance"] == "Present")
        absent = total - present
        rate = round((present / total * 100) if total > 0 else 0, 1)
        
        return jsonify({
            'records': records,
            'summary': {
                'total': total,
                'present': present,
                'absent': absent,
                'rate': rate
            }
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Preview compilation failed: {str(e)}'}), 500


@attendance_bp.route('/save-run', methods=['POST'])
@token_required
def save_run(current_user):
    """Saves compiled run to SQL history database using raw SQL."""
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    summary = data.get('summary', {})
    
    if not date_str or not records:
        return jsonify({'message': 'Date and student records are required.'}), 400
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Delete duplicate run if exists for this user/date
        cursor.execute(f"SELECT id FROM history_{current_user.username} WHERE attendance_date = ?", (date_str,))
        existing = cursor.fetchone()
        if existing:
            run_id = existing['id']
            cursor.execute(f"DELETE FROM attendance_{current_user.username} WHERE run_id = ?", (run_id,))
            cursor.execute(f"DELETE FROM history_{current_user.username} WHERE id = ?", (run_id,))
            
        total = len(records)
        present = sum(1 for r in records if r["attendance"] == "Present")
        absent = total - present
        excel_path = str(get_user_output_dir(current_user.username) / f"Academic_Attendance_{date_str}.xlsx")
        sync_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        cursor.execute(f"""
            INSERT INTO history_{current_user.username}
            (attendance_date, sync_time, total_students, present_count, absent_count, excel_file_path)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (date_str, sync_time, total, present, absent, excel_path))
        
        new_run_id = cursor.lastrowid
        
        # Save details
        for r in records:
            cursor.execute(f"""
                INSERT INTO attendance_{current_user.username}
                (run_id, roll_number, enrollment_number, student_name, attendance)
                VALUES (?, ?, ?, ?, ?)
            """, (new_run_id, r["roll_number"], r["enrollment_number"], r["student_name"], r["attendance"]))
            
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Successfully saved run to local database.'
        }), 201
    except Exception as e:
        return jsonify({'message': f'Failed to save run: {str(e)}'}), 500


@attendance_bp.route('/export-excel', methods=['POST'])
@token_required
def export_excel(current_user):
    """Generates styled Excel binary file stream for download."""
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    
    if not date_str or not records:
        return jsonify({'message': 'Date and student records are required.'}), 400
        
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Academic Attendance"
        ws.views.sheetView[0].showGridLines = True
        
        font_name = "Segoe UI"
        title_font = Font(name=font_name, size=16, bold=True, color="1E3D59")
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=11)
        
        header_fill = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
        present_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        absent_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Header title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"ACADEMIC ATTENDANCE AUTOMATION - {date_str}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers Row
        headers = ["Roll No", "Enrollment No", "Student Name", "Attendance"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Data Rows
        for r_idx, r in enumerate(records, 4):
            ws.cell(row=r_idx, column=1, value=r["roll_number"]).font = data_font
            ws.cell(row=r_idx, column=2, value=r["enrollment_number"]).font = data_font
            ws.cell(row=r_idx, column=3, value=r["student_name"]).font = data_font
            
            status_cell = ws.cell(row=r_idx, column=4, value=r["attendance"])
            status_cell.font = data_font
            if r["attendance"] == "Present":
                status_cell.fill = present_fill
            else:
                status_cell.fill = absent_fill
                
        # Send back in-memory bytes stream
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return send_file(
            file_stream,
            as_attachment=True,
            download_name=f"Academic_Attendance_{date_str}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({'message': f'Failed to generate Excel: {str(e)}'}), 500


@attendance_bp.route('/export-pdf', methods=['POST'])
@token_required
def export_pdf(current_user):
    """Generates styled PDF binary file stream for download."""
    data = request.get_json() or {}
    date_str = data.get('date')
    records = data.get('records', [])
    
    if not date_str or not records:
        return jsonify({'message': 'Date and student records are required.'}), 400
        
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='TitleStyle',
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#1E3D59'),
            alignment=1, # Center
            spaceAfter=15
        )
        
        story = []
        story.append(Paragraph(f"Academic Attendance Report - {date_str}", title_style))
        story.append(Spacer(1, 10))
        
        # Table data
        table_data = [["Roll No", "Enrollment No", "Student Name", "Attendance"]]
        for r in records:
            table_data.append([
                r["roll_number"],
                r["enrollment_number"],
                r["student_name"],
                r["attendance"]
            ])
            
        t = Table(table_data, colWidths=[80, 100, 240, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3D59')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ]))
        
        # Color specific rows depending on Present/Absent
        for i, row in enumerate(records, 1):
            color = colors.HexColor('#E2F0D9') if row["attendance"] == "Present" else colors.HexColor('#FCE4D6')
            t.setStyle(TableStyle([
                ('BACKGROUND', (3, i), (3, i), color),
                ('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#385723') if row["attendance"] == "Present" else colors.HexColor('#C00000')),
                ('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'),
            ]))
            
        story.append(t)
        doc.build(story)
        
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Academic_Attendance_{date_str}.pdf",
            mimetype="application/pdf"
        )
    except Exception as e:
        return jsonify({'message': f'Failed to generate PDF: {str(e)}'}), 500


@attendance_bp.route('/history', methods=['GET'])
@token_required
def get_history(current_user):
    """Retrieves paginated and searchable history of attendance runs for the user using raw SQL."""
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    search = request.args.get('search', '').strip()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        sql = f"SELECT * FROM history_{current_user.username}"
        params = []
        if search:
            sql += " WHERE attendance_date LIKE ?"
            params.append(f"%{search}%")
            
        cursor.execute(f"SELECT COUNT(*) FROM ({sql})", params)
        total_matching = cursor.fetchone()[0]
        
        offset = (page - 1) * limit
        sql += " ORDER BY attendance_date DESC LIMIT ? OFFSET ?"
        cursor.execute(sql, params + [limit, offset])
        rows = cursor.fetchall()
        conn.close()
        
        runs = []
        for r in rows:
            pct = (r['present_count'] / r['total_students'] * 100) if r['total_students'] > 0 else 0
            runs.append({
                "id": r['id'],
                "attendance_date": r['attendance_date'],
                "sync_time": r['sync_time'],
                "total_students": r['total_students'],
                "present_count": r['present_count'],
                "absent_count": r['absent_count'],
                "rate": round(pct, 1)
            })
            
        total_pages = (total_matching + limit - 1) // limit if total_matching > 0 else 1
        return jsonify({
            "runs": runs,
            "total_pages": total_pages,
            "total_runs": total_matching,
            "current_page": page
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve history: {str(e)}'}), 500


@attendance_bp.route('/history/<int:run_id>', methods=['GET'])
@token_required
def get_run_details(current_user, run_id):
    """Retrieves list of student attendance rows under a run, supporting search/status filters using raw SQL."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM history_{current_user.username} WHERE id = ?", (run_id,))
        run = cursor.fetchone()
        if not run:
            conn.close()
            return jsonify({'message': 'Run not found.'}), 404
            
        search_val = request.args.get('search', '').strip().lower()
        status_val = request.args.get('status', '').strip().lower()
        
        sql = f"SELECT * FROM attendance_{current_user.username} WHERE run_id = ?"
        params = [run_id]
        if status_val:
            sql += " AND LOWER(attendance) = ?"
            params.append(status_val)
            
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        conn.close()
        
        filtered_details = []
        for d in rows:
            roll = d['roll_number'] or ""
            enroll = d['enrollment_number'] or ""
            name = d['student_name'] or ""
            status = d['attendance'] or ""
            
            name_matches = (search_val == '') or (
                search_val in name.lower() or 
                search_val in roll.lower() or 
                search_val in enroll.lower()
            )
            
            if name_matches:
                filtered_details.append({
                    "id": d["id"],
                    "roll_number": roll,
                    "enrollment_number": enroll,
                    "student_name": name,
                    "attendance": status
                })
                
        pct = (run['present_count'] / run['total_students'] * 100) if run['total_students'] > 0 else 0
        return jsonify({
            "run": {
                "id": run['id'],
                "attendance_date": run['attendance_date'],
                "total_students": run['total_students'],
                "present_count": run['present_count'],
                "absent_count": run['absent_count'],
                "rate": round(pct, 1)
            },
            "details": filtered_details
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve details: {str(e)}'}), 500


@attendance_bp.route('/history/<int:run_id>', methods=['DELETE'])
@token_required
def delete_run(current_user, run_id):
    """Deletes an attendance run and cascades to all detailed student marks using raw SQL."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT id FROM history_{current_user.username} WHERE id = ?", (run_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'message': 'Run not found.'}), 404
            
        cursor.execute(f"DELETE FROM attendance_{current_user.username} WHERE run_id = ?", (run_id,))
        cursor.execute(f"DELETE FROM history_{current_user.username} WHERE id = ?", (run_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Attendance run deleted successfully.'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to delete run: {str(e)}'}), 500


@attendance_bp.route('/reports/stats', methods=['GET'])
@token_required
def get_reports_stats(current_user):
    """Retrieves options (dates, month labels, run summaries) to configure reports generation using raw SQL."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM history_{current_user.username} ORDER BY attendance_date DESC")
        rows = cursor.fetchall()
        conn.close()
        
        dates = []
        months_seen = set()
        months_options = []
        
        total_runs = len(rows)
        total_present = 0
        total_students = 0
        
        for r in rows:
            date_str = r['attendance_date']
            dates.append(date_str)
            
            # Aggregate stats
            total_present += r['present_count']
            total_students += r['total_students']
            
            # Parse months
            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                year = dt.year
                month = dt.month
                month_label = dt.strftime("%B %Y")
                month_key = f"{year}-{month}"
                
                if month_key not in months_seen:
                    months_seen.add(month_key)
                    months_options.append({
                        "year": year,
                        "month": month,
                        "label": month_label
                    })
            except Exception:
                pass
                
        avg_rate = round((total_present / total_students * 100) if total_students > 0 else 0, 1)
        
        return jsonify({
            "dates": dates,
            "months": months_options,
            "summary": {
                "total_runs": total_runs,
                "average_rate": avg_rate
            }
        }), 200
    except Exception as e:
        return jsonify({'message': f'Failed to retrieve report metadata: {str(e)}'}), 500


@attendance_bp.route('/reports/generate', methods=['POST'])
@token_required
def generate_report_file(current_user):
    """Compiles daily, monthly, or custom matrix reports as Excel or PDF downloads using raw SQL."""
    from app.utils.report_generator import compile_runs_list, build_matrix_report_excel, build_matrix_report_pdf

    data = request.get_json() or {}
    report_type = data.get('report_type', 'daily') # 'daily', 'monthly', 'custom'
    fmt = data.get('format', 'excel') # 'excel', 'pdf'
    
    try:
        if report_type == 'daily':
            date_str = data.get('date')
            if not date_str:
                return jsonify({'message': 'Date parameter is required for Daily reports.'}), 400
                
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM history_{current_user.username} WHERE attendance_date = ?", (date_str,))
            run = cursor.fetchone()
            conn.close()
            
            if not run:
                return jsonify({'message': f'No attendance run found for date {date_str}.'}), 404
                
            # Reuses existing single day export endpoints internally
            if fmt == 'pdf':
                return export_pdf(current_user=current_user)
            else:
                return export_excel(current_user=current_user)
                
        elif report_type == 'monthly':
            year = data.get('year')
            month = data.get('month')
            if not year or not month:
                return jsonify({'message': 'Year and Month parameters are required.'}), 400
                
            runs = compile_runs_list(current_user.username, year=int(year), month=int(month))
            if not runs:
                return jsonify({'message': 'No attendance history records found for the selected month.'}), 404
                
            month_name = datetime.date(int(year), int(month), 1).strftime("%B %Y")
            title_text = f"MONTHLY ATTENDANCE REPORT - {month_name.upper()}"
            
            if fmt == 'pdf':
                file_stream = build_matrix_report_pdf(current_user.username, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Monthly_Attendance_Report_{year}_{month}.pdf",
                    mimetype="application/pdf"
                )
            else:
                file_stream = build_matrix_report_excel(current_user.username, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Monthly_Attendance_Report_{year}_{month}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        elif report_type == 'custom':
            start_date_str = data.get('start_date')
            end_date_str = data.get('end_date')
            if not start_date_str or not end_date_str:
                return jsonify({'message': 'Start and End dates are required.'}), 400
                
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            
            runs = compile_runs_list(current_user.username, start_date=start_date, end_date=end_date)
            if not runs:
                return jsonify({'message': 'No attendance history records found for the specified range.'}), 404
                
            title_text = f"ATTENDANCE SUMMARY REPORT ({start_date_str} TO {end_date_str})"
            
            if fmt == 'pdf':
                file_stream = build_matrix_report_pdf(current_user.username, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Attendance_Report_{start_date_str}_to_{end_date_str}.pdf",
                    mimetype="application/pdf"
                )
            else:
                file_stream = build_matrix_report_excel(current_user.username, title_text, runs)
                return send_file(
                    file_stream,
                    as_attachment=True,
                    download_name=f"Attendance_Report_{start_date_str}_to_{end_date_str}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            return jsonify({'message': 'Unsupported report type.'}), 400
            
    except Exception as e:
        return jsonify({'message': f'Report generation failed: {str(e)}'}), 500
