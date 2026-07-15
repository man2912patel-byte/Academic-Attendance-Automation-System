import os
import logging
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils import parse_date_string

logger = logging.getLogger("attendance")

def get_date_candidates(s):
    """
    Intelligently parses date strings in multiple formats (DD-MM-YYYY, MM-DD-YYYY, etc.)
    and returns a list of unique datetime.date objects that this string could represent.
    """
    if not s or not isinstance(s, str):
        return []
        
    s = s.strip()
    # Normalize separators to hyphen
    s_norm = s.replace("/", "-").replace(".", "-").replace(" ", "-")
    
    # Standard format attempts
    fmts = [
        # 4-digit Year
        "%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%d",
        "%d-%b-%Y", "%b-%d-%Y",
        "%d-%B-%Y", "%B-%d-%Y",
        
        # 2-digit Year
        "%d-%m-%y", "%m-%d-%y", "%y-%m-%d",
        "%d-%b-%y", "%b-%d-%y",
    ]
    
    candidates = []
    
    # First, try formats on the normalized string
    for fmt in fmts:
        try:
            dt = datetime.datetime.strptime(s_norm, fmt)
            if 2000 <= dt.year <= 2100:
                d = dt.date()
                if d not in candidates:
                    candidates.append(d)
        except ValueError:
            pass

    # Try original string with spaces (for formats like "14 Jul 2026")
    for fmt in fmts:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            if 2000 <= dt.year <= 2100:
                d = dt.date()
                if d not in candidates:
                    candidates.append(d)
        except ValueError:
            pass

    return candidates

def resolve_date_sequence(raw_date_strings):
    """
    Given a list of raw date strings, parses them and resolves any ambiguities
    by finding a non-decreasing sequence of dates that minimizes the total span
    (end_date - start_date) and matches the chronological flow.
    """
    if not raw_date_strings:
        return []

    # 1. Get candidates for each string
    candidates_list = []
    for s in raw_date_strings:
        cands = get_date_candidates(s)
        if not cands:
            logger.warning(f"Could not parse any date candidates from string: '{s}'")
            cands = [datetime.date.today()]
        candidates_list.append(cands)

    n = len(candidates_list)

    # 2. Find a non-decreasing sequence using backtracking
    valid_sequences = []

    def backtrack(idx, current_seq):
        if len(valid_sequences) >= 1000:
            return
            
        if idx == n:
            valid_sequences.append(list(current_seq))
            return
        
        last_date = current_seq[-1] if current_seq else None
        
        for cand in sorted(candidates_list[idx]):
            if last_date is None or cand >= last_date:
                current_seq.append(cand)
                backtrack(idx + 1, current_seq)
                current_seq.pop()

    backtrack(0, [])

    if not valid_sequences:
        logger.error("Could not find any valid non-decreasing sequence of dates. Using fallbacks.")
        return [cands[0] for cands in candidates_list]

    # Find the sequence with the minimum span (end_date - start_date)
    best_seq = None
    min_span = None

    for seq in valid_sequences:
        span = (seq[-1] - seq[0]).days
        if min_span is None or span < min_span:
            min_span = span
            best_seq = seq

    return best_seq

class AttendanceManager:
    def __init__(self, config):
        self.config = config

    def parse_sheets_data(self, mft_raw, marquee_raw):
        """
        Processes raw MFT student list and Marquee attendance list.
        Detects columns dynamically by headers, performs matching,
        and extracts available dates.
        """
        # --- 1. Parse MFT Student List ---
        if not mft_raw or len(mft_raw) < 1:
            raise ValueError("MFT Student list is empty or invalid.")
            
        mft_headers = [str(h).strip().lower() for h in mft_raw[0].keys()]
        
        # Find column mappings in MFT list
        email_key = None
        roll_key = None
        
        for k in mft_raw[0].keys():
            k_low = k.strip().lower()
            if "email" in k_low or "mail" in k_low:
                email_key = k
            elif "rhn" in k_low or "roll" in k_low or "id" in k_low:
                roll_key = k

        # Fallback to first/second columns if headers not clearly identified
        if not email_key and mft_raw[0]:
            email_key = list(mft_raw[0].keys())[min(1, len(mft_raw[0])-1)]
        if not roll_key and mft_raw[0]:
            roll_key = list(mft_raw[0].keys())[0]

        logger.info(f"MFT keys mapped - Email: '{email_key}', Roll/RHN: '{roll_key}'")

        # Convert MFT students to standard dictionaries
        mft_students = []
        for row in mft_raw:
            email = str(row.get(email_key, "")).strip()
            roll = str(row.get(roll_key, "")).strip()
            
            # Extract enrollment number from email if possible (username part before @)
            enrollment = ""
            if email and "@" in email:
                username = email.split("@")[0]
                if username.isdigit():
                    enrollment = username
            
            mft_students.append({
                "raw_email": email,
                "roll_no": roll,
                "extracted_enrollment": enrollment
            })

        logger.info(f"Processed {len(mft_students)} MFT students.")

        # --- 2. Parse Marquee Attendance Sheet ---
        if not marquee_raw or len(marquee_raw) < 2:
            raise ValueError("Marquee attendance sheet must have at least 2 header rows.")

        row_0 = [str(cell).strip() for cell in marquee_raw[0]]
        row_1 = [str(cell).strip() for cell in marquee_raw[1]]
        data_rows = marquee_raw[2:]

        # Find static column indices dynamically
        enroll_col_idx = None
        name_col_idx = None
        email_col_idx = None
        
        for idx, col_name in enumerate(row_0):
            col_name_low = col_name.lower()
            if "enrollment" in col_name_low or "enroll" in col_name_low:
                enroll_col_idx = idx
            elif "name" in col_name_low or "student" in col_name_low:
                name_col_idx = idx
            elif "mail" in col_name_low or "email" in col_name_low:
                email_col_idx = idx

        # Fallback indexes if not detected
        if enroll_col_idx is None:
            enroll_col_idx = 6 # Row 0 index 6 in sample
        if name_col_idx is None:
            name_col_idx = 7   # Row 0 index 7 in sample
        if email_col_idx is None:
            email_col_idx = 14  # Row 0 index 14 in sample

        logger.info(f"Marquee columns mapped - Enrollment: Index {enroll_col_idx}, Name: Index {name_col_idx}, Email: Index {email_col_idx}")

        # Scan columns starting from 15 (dates and sessions)
        raw_headers_list = []
        last_raw_date = ""

        for idx in range(min(len(row_0), len(row_1))):
            if idx < 15:
                continue
                
            cell_0 = row_0[idx].strip()
            cell_1 = row_1[idx].strip()
            
            # Forward fill the date header
            if cell_0:
                if get_date_candidates(cell_0):
                    last_raw_date = cell_0
                else:
                    last_raw_date = ""
            
            if last_raw_date:
                raw_headers_list.append((idx, last_raw_date, cell_1))

        # Resolve raw date strings to concrete, correct date objects
        raw_date_strings = [item[1] for item in raw_headers_list]
        resolved_dates = resolve_date_sequence(raw_date_strings)

        date_map = {} # date_obj -> [(col_idx, session_name)]
        for i, (col_idx, _, session_name) in enumerate(raw_headers_list):
            if i < len(resolved_dates):
                date_obj = resolved_dates[i]
                session_name = session_name if session_name else f"Session {col_idx}"
                if date_obj not in date_map:
                    date_map[date_obj] = []
                date_map[date_obj].append((col_idx, session_name))

        logger.info(f"Detected {len(date_map)} unique attendance dates from Marquee sheet.")
        
        # Build Marquee Student Directory for matching
        marquee_students = []
        for r_idx, row in enumerate(data_rows):
            if len(row) <= max(enroll_col_idx, name_col_idx, email_col_idx):
                continue
                
            enroll = str(row[enroll_col_idx]).strip()
            name = str(row[name_col_idx]).strip()
            email = str(row[email_col_idx]).strip()
            
            if not enroll and not name:
                continue
                
            marquee_students.append({
                "row_idx": r_idx,
                "enrollment": enroll,
                "name": name,
                "email": email,
                "raw_row": row
            })

        logger.info(f"Parsed {len(marquee_students)} student rows from Marquee list.")

        return mft_students, marquee_students, date_map

    def match_students(self, mft_students, marquee_students):
        """
        Matches MFT students with Marquee students using multi-tier key matching.
        """
        matched_records = []
        mismatched_mft = []

        # Build lookup tables for Marquee students
        marquee_by_enroll = {}
        marquee_by_email = {}
        marquee_by_name = {}

        for m_student in marquee_students:
            enroll_clean = m_student["enrollment"].replace(" ", "").strip()
            email_clean = m_student["email"].lower().strip()
            name_clean = m_student["name"].lower().strip()

            if enroll_clean:
                marquee_by_enroll[enroll_clean] = m_student
            if email_clean:
                marquee_by_email[email_clean] = m_student
            if name_clean:
                marquee_by_name[name_clean] = m_student

        # Perform matching
        for mft in mft_students:
            mft_enroll = mft["extracted_enrollment"].strip()
            mft_roll = mft["roll_no"].strip()
            mft_email = mft["raw_email"].lower().strip()

            matched_marquee = None

            # Tier 1: Match by enrollment number (extracted from MFT Email or roll number)
            if mft_enroll and mft_enroll in marquee_by_enroll:
                matched_marquee = marquee_by_enroll[mft_enroll]
            elif mft_roll and mft_roll in marquee_by_enroll:
                matched_marquee = marquee_by_enroll[mft_roll]

            # Tier 2: Match by exact Email
            if not matched_marquee and mft_email:
                if mft_email in marquee_by_email:
                    matched_marquee = marquee_by_email[mft_email]

            # Tier 3: Match by Roll Number digits in enrollment (partial match)
            if not matched_marquee:
                roll_digits = "".join(filter(str.isdigit, mft_roll))
                if roll_digits and len(roll_digits) >= 4:
                    for enroll_key, m_student in marquee_by_enroll.items():
                        if enroll_key.endswith(roll_digits):
                            matched_marquee = m_student
                            break
            
            if matched_marquee:
                matched_records.append({
                    "mft_student": mft,
                    "marquee_student": matched_marquee
                })
            else:
                mismatched_mft.append(mft)

        logger.info(f"Matched {len(matched_records)} students. {len(mismatched_mft)} MFT students could not be matched.")
        return matched_records, mismatched_mft

    def compute_attendance(self, matched_records, mismatched_mft, date_obj, session_cols, date_map, mode="Combined (Any)"):
        """
        Computes Present/Absent for matched students for the specified date and session.
        Modes:
          - "Combined (Any)": Present if Present in any of the sessions.
          - "Combined (All)": Present only if Present in all sessions.
          - Specific session name (e.g. "Session 01")
        """
        results = []
        
        if not session_cols:
            raise ValueError(f"No session columns found for date {date_obj}")

        # Compute attendance for matched records
        for record in matched_records:
            mft = record["mft_student"]
            marquee = record["marquee_student"]
            raw_row = marquee["raw_row"]

            # Collect statuses for each session column
            statuses = []
            for col_idx, sess_name in session_cols:
                if col_idx < len(raw_row):
                    val = str(raw_row[col_idx]).strip().lower()
                    statuses.append((sess_name, val))
                else:
                    statuses.append((sess_name, "absent"))

            # Determine final status
            final_status = "Absent"
            
            if mode == "Combined (Any)":
                is_present = any("present" in s[1] for s in statuses)
                final_status = "Present" if is_present else "Absent"
            elif mode == "Combined (All)":
                is_present = all("present" in s[1] for s in statuses)
                final_status = "Present" if is_present else "Absent"
            else:
                val = "absent"
                for sess_name, s_val in statuses:
                    if sess_name == mode:
                        val = s_val
                        break
                final_status = "Present" if "present" in val else "Absent"

            results.append({
                "roll_number": mft["roll_no"],
                "enrollment_number": marquee["enrollment"],
                "student_name": marquee["name"],
                "attendance": final_status,
                "date": date_obj.strftime("%Y-%m-%d"),
                "time_generated": datetime.datetime.now().strftime("%H:%M:%S")
            })

        # Process unmatched MFT students as Absent (Not Found in Marquee)
        for mft in mismatched_mft:
            results.append({
                "roll_number": mft["roll_no"],
                "enrollment_number": mft["extracted_enrollment"] or "N/A",
                "student_name": "Unknown Student (New/Not Synced)",
                "attendance": "Absent",
                "date": date_obj.strftime("%Y-%m-%d"),
                "time_generated": datetime.datetime.now().strftime("%H:%M:%S")
            })

        return results

    def save_to_excel(self, records, date_obj, output_dir=None):
        """
        Generates and saves the styled Academic_Attendance_YYYY-MM-DD.xlsx sheet.
        """
        if output_dir is None:
            output_dir = Path(self.config.output_dir)
        else:
            output_dir = Path(output_dir)
            
        output_dir.mkdir(parents=True, exist_ok=True)
        
        date_str = date_obj.strftime("%Y-%m-%d")
        filename = output_dir / f"Academic_Attendance_{date_str}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Academic Attendance"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_name = "Segoe UI"
        title_font = Font(name=font_name, size=16, bold=True, color="1E3D59")
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=11)
        meta_font = Font(name=font_name, size=10, italic=True, color="555555")
        
        header_fill = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
        zebra_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        present_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        present_font = Font(name=font_name, size=11, bold=True, color="385723")
        
        absent_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        absent_font = Font(name=font_name, size=11, bold=True, color="C00000")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # 1. Add Title Banner
        ws.merge_cells("A1:F1")
        ws["A1"] = "MFT ACADEMIC ATTENDANCE AUTOMATION SYSTEM"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # 2. Add Meta Information
        now_time = datetime.datetime.now().strftime("%I:%M:%S %p")
        ws["A2"] = f"Attendance Date: {date_str}"
        ws["A2"].font = Font(name=font_name, size=11, bold=True)
        ws["F2"] = f"Generated: {now_time}"
        ws["F2"].font = meta_font
        ws["F2"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[2].height = 20
        
        # Spacer row
        ws.row_dimensions[3].height = 10
        
        # 3. Add Table Headers
        headers = ["Roll Number", "Enrollment Number", "Student Name", "Attendance", "Date", "Time Generated"]
        ws.append([])
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[5].height = 25
 
        sorted_records = sorted(records, key=lambda x: x["roll_number"])

        # 4. Add Data Rows
        start_row = 6
        for r_idx, rec in enumerate(sorted_records, start_row):
            ws.cell(row=r_idx, column=1, value=rec["roll_number"])
            ws.cell(row=r_idx, column=2, value=rec["enrollment_number"])
            ws.cell(row=r_idx, column=3, value=rec["student_name"])
            
            att_cell = ws.cell(row=r_idx, column=4, value=rec["attendance"])
            
            ws.cell(row=r_idx, column=5, value=rec["date"])
            ws.cell(row=r_idx, column=6, value=rec["time_generated"])
            
            is_zebra = (r_idx % 2 == 0)
            row_fill = zebra_fill if is_zebra else white_fill
            
            for col_num in range(1, 7):
                c = ws.cell(row=r_idx, column=col_num)
                c.font = data_font
                c.fill = row_fill
                c.border = thin_border
                
                if col_num in [1, 2, 5, 6]:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num == 3:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                elif col_num == 4:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            
            if rec["attendance"] == "Present":
                att_cell.fill = present_fill
                att_cell.font = present_font
            else:
                att_cell.fill = absent_fill
                att_cell.font = absent_font
                
            ws.row_dimensions[r_idx].height = 20

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col[4:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(filename)
        logger.info(f"Excel report saved successfully to {filename}")
        return filename
